from django.shortcuts import render_to_response, redirect, render, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse
from django.core.files import File
from django.core.exceptions import ValidationError
from django.core.context_processors import csrf
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.template import Context
from django.template.response import TemplateResponse
from django.template import RequestContext
from django.utils.translation import ugettext, ugettext_lazy as _
from django.utils.encoding import smart_unicode, force_unicode
from django.utils import simplejson
from django.conf import settings

from eventswat.models import *
from eventswat.forms import *
from events.models import *
from usermanagement	.models import Userprofile
from usermanagement.forms import UserCreationForm, UserLoginForm
from events.extra import JSONResponse
from postevent.models import Postevent, Organizer, CampusCollege, CampusDepartment
from reviews.models import *
from reviews.forms import *
from postbanner.models import *
from payu.models import *
from eventswat.util import format_redirect_url
from templated_email import send_templated_mail
from forms import UploadFileForm
from django.utils.timezone import utc
import random
import string
import datetime
import time
import openpyxl

import json
# Import for google calendar
# from __future__ import print_function
# import httplib2
# import os
# from apiclient import discovery
# import oauth2client
# from oauth2client import client
# from oauth2client import tools
from datetime import datetime, timedelta
from json import dumps


def default(obj):
    """Default JSON serializer."""
    import calendar
    import datetime

    if isinstance(obj, datetime.datetime):
        if obj.utcoffset() is not None:
            obj = obj - obj.utcoffset()
        millis = int(
            calendar.timegm(obj.timetuple()) * 1000 +
            obj.microsecond / 1000
        )
        return millis
    raise TypeError('Not sure how to serialize %s' % (obj,))


@csrf_exempt
def home(request):
    registeration_form = UserCreationForm()
    login_form = UserLoginForm()
    feedback_form = WebsiteFeedbackForm()
    if request.user.is_superuser:
        logout(request)
        return HttpResponseRedirect('/')
    return render_to_response("index_v2.html", {'registeration_form': registeration_form, 'login_form': login_form, 'feedback_form': feedback_form}, context_instance=RequestContext(request))


def about(request):
    return render_to_response("about-us.html", context_instance=RequestContext(request))


def privacy_and_policy(request):
    return render_to_response("privacy.html", context_instance=RequestContext(request))


def terms_and_conditions(request):
    return render_to_response("terms_and_conditions.html", context_instance=RequestContext(request))


def faqs(request):
    return render_to_response("faqs.html", context_instance=RequestContext(request))


def start(request):
    return render_to_response('index_v2.html', context_instance=RequestContext(request))


def logout_view(request):
    logout(request)
    response = HttpResponseRedirect("/")
    return response

# comments implemented by priya


@csrf_exempt

def details(request,id=None):
	# try:
	postevent=Postevent.objects.get(pk=id)
	userprofile = Userprofile()
	img=str(postevent.event_poster).split(',')
	photo=img[0]
	photos=[n for n in str(postevent.event_poster).split(',')]
	organizer=Organizer.objects.filter(postevent__id=postevent.id)
	registeration_form = UserCreationForm()
	login_form = UserLoginForm()
	comment_form = CommentForm()
	comment_tree=Comment.objects.filter(postevent_id=postevent.id).order_by('-path')
	# related_events = Postevent.objects.filter(event_category = postevent.event_category, event_subcategory=postevent.event_subcategory, city=postevent.city)
	return render_to_response("company-profile.html",{'comment_tree':comment_tree,'events':postevent,'organizer':organizer,'photos':photos,'photo':photo, 'registeration_form':registeration_form, 'login_form':login_form, 'comment_form':comment_form, 'userprofile':userprofile}, context_instance=RequestContext(request))
	# response.set_cookie( 'content', temp.content )
	# response.set_cookie( 'rating', temp.rating )
	
	# except:
	#     return render_to_response("company-profile.html",{'message':'Sorry for inconvenience.Some thing went to wrong'}, context_instance=RequestContext(request))

def banner(request):
	return render_to_response("uploadbanner.html",context_instance=RequestContext(request))

# login and registration implemanted by ramya
@csrf_exempt
def user_login(request):
	import json
	if request.user.is_superuser:
		logout(request)
		return HttpResponseRedirect('/')
	logout(request)
	error = {}
	if request.method == 'POST':
		print "post1"
		login_form = UserLoginForm(request.POST)
		
		login_email = login_form.cleaned_data['login_email']
		print 'email form form', login_email
		login_password = login_form.cleaned_data['login_password']
		print 'Login_password', login_password
		context = {}
		if not User.objects.filter(email=login_email).exists():
			
			error['email_exists'] = True
			response = HttpResponse(json.dumps(error, ensure_ascii=False),mimetype='application/json')
			return response

		if login_form.is_valid():
			user = User.objects.get(email=login_email)
			user.backend='django.contrib.auth.backends.ModelBackend'
			if user is not None:
				if user.check_password(login_password):
					if user.is_active:
						login(request, user)
						response = HttpResponseRedirect(request.POST.get('next'))
				else:
					error['password'] = True
					
					response = HttpResponse(json.dumps(error, ensure_ascii=False),mimetype='application/json')
					return response
	else:
		print 'this else'
		login_form = UserLoginForm()
	return render_to_response('index_v2.html', {'login_form':login_form}, context_instance=RequestContext(request))

@csrf_protect
def register(request):
    registered = False
    user = User()
    userprofile = Userprofile()
    if request.method == 'POST':       
        registeration_form = UserCreationForm(request.POST)       
        username = registeration_form.data.get('username')       
        email = registeration_form.cleaned_data['email']
   
        try:
            error = {}
            if Userprofile.objects.filter(username=username).exists():              
                error['username_exists'] = ugettext('Username already exists')
                raise ValidationError(error['username_exists'], 1)
            if Userprofile.objects.filter(email=email).exists():
                error['email_exists'] = ugettext('Email already exists')
                raise ValidationError(error['email_exists'], 2)

        except ValidationError as e:
            print 'except'
            messages.add_message(request, messages.ERROR, e.messages[-1])
            redirect_path = "/"
            query_string = 'rst=%d' % e.code
            redirect_url = format_redirect_url(redirect_path, query_string)
            return HttpResponseRedirect(redirect_url)

        if registeration_form.is_valid():
            print 'is_valid'
            print email, 'userprofile.email'
            userprofile.is_active = True
            userprofile = registeration_form.save()
            send_templated_mail(
                template_name='welcome',
                subject='Welcome Evewat',
                from_email='eventswat@gmail.com',
                recipient_list=[email],
                context={
                    'userprofile': username,
                },
            )
            registered = True
            registered_user = Userprofile.objects.get(email=email)
            print 'user after reg', registered_user
            registered_user.backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, registered_user)
            return HttpResponseRedirect('/start/?user_id=' + str(registered_user.id))

        elif userprofile.id is None:
            print 'userprofile is none'
            return HttpResponseRedirect('/')

    else:
        print "else"
        registeration_form = UserCreationForm()
        user_id = userprofile.id
    return render_to_response('index_v2.html', {'user_id': user_id, 'registeration_form': registeration_form}, context_instance=RequestContext(request))


@csrf_exempt
@login_required(login_url='/?lst=1')
def post_event(request):
    try:
        eventscategory = EventsCategory.objects.all()
        eventssubcategory = EventsSubCategory.objects.all()
        state = City.objects.values_list('state', flat=True)
        # premium=PremiumPriceInfo.objects.all()
        return render_to_response("post_event.html", {'eventssubcategory': eventssubcategory, 'eventscategory': eventscategory, 'state': list(set(state))}, context_instance=RequestContext(request))
    except:
        return render_to_response("post_event.html", {'message': 'Sorry for inconvenience.Some thing went to wrong'}, context_instance=RequestContext(request))


def submit_event_v2(request):
    try:
        if request.method == "POST":
            postevent = Postevent()
            postevent.user_name = request.POST['name']
            postevent.user_email = request.POST['email']
            postevent.user_mobile = request.POST.get('mobile', '0')
            postevent.event_title = request.POST.get('eventtitle', '')
            startdate = request.POST.get('startdate', '')
            date, month, year = startdate.split('-')
            postevent.event_startdate_time = year + '-' + month + '-' + date
            enddate = request.POST.get('enddate', '')
            date, month, year = enddate.split('-')
            postevent.event_enddate_time = year + '-' + month + '-' + date
            postevent_category = EventsCategory.objects.get(
                id=request.POST.get('category', ''))
            postevent.event_category = postevent_category
            postevent_subcategory = EventsSubCategory.objects.get(
                id=request.POST.get('eventtype', ''))
            postevent.event_subcategory = postevent_subcategory
            postevent.event_description = request.POST.get(
                'eventdescription', '')
            postevent.venue = request.POST.get('address', '')
            postevent.state = request.POST.get('state', '')
            postevent_city = request.POST.get('city', '')
            postevent.city = postevent_city
            postevent_college = request.POST.get('college', '')
            postevent.college = postevent_college
            postevent.department = request.POST.get('dept', '')
            postevent_poster = request.FILES.getlist('poster[]')

            def handle_uploaded_file(f):
                postevent_poster = open(
                    settings.MEDIA_ROOT + '/events/' + '%s' % f.name, 'wb+')
                for chunk in f.chunks():
                    postevent_poster.write(chunk)
                postevent_poster.close()
            photosgroup = ''
            count = len(postevent_poster)

            if count:
                for uploaded_file in postevent_poster:
                    count = count - 1
                    handle_uploaded_file(uploaded_file)
                    if count == 0:
                        photosgroup = photosgroup + \
                            '/events/' + str(uploaded_file)
                    else:
                        photosgroup = photosgroup + \
                            '/events/' + str(uploaded_file) + ','
                postevent.event_poster = photosgroup
            else:
                postevent.event_poster = '/events/img/logo_150.png'
            # if request.POST.get('plan')!='0':
            #     postevent.payment=request.POST.get('plan')
            postevent.save()
            organizer = Organizer()
            post = Postevent.objects.order_by('-pk')[0]
            organizer.postevent = postevent
            organizer.organizer_name = request.POST.get('organizer_name', '')
            organizer.organizer_mobile = request.POST.getlist(
                'organizer_mobile[]', '')
            mobile_number = ''
            count = len(organizer.organizer_mobile)
            if count:
                for number in organizer.organizer_mobile:
                    count = count - 1
                    if count == 0:
                        mobile_number = mobile_number + number
                    else:
                        mobile_number = mobile_number + number + ','
            organizer.organizer_mobile = mobile_number
            organizer.organizer_email = request.POST.get('organizer_email', '')
            organizer.save()
            send_templated_mail(
                template_name='post_event',
                subject='Post Event',
                from_email='eventswat@gmail.com',
                recipient_list=[postevent.user_email],
                context={

                    'user': postevent.user_name,

                },
            )
            message = "Your data succesfully submitted"

        # user_amount=request.POST.get('plan')
        # if user_amount!='0':
        #     return HttpResponseRedirect('/payment_event/')
        # elif user_amount=='0':
        #     response=render_to_response("post_event.html",{'message':message}, context_instance=RequestContext(request))
        # else:
            response = render_to_response("post_event.html", {
                                          'message': message}, context_instance=RequestContext(request))
            response.delete_cookie('eventtitle')
            response.delete_cookie('startdate')
            response.delete_cookie('enddate')
        # response.delete_cookie('plan')
            response.delete_cookie('category_name')
            response.delete_cookie('eventtype_name')
            response.delete_cookie('eventtype')
            response.delete_cookie('category')
            response.delete_cookie('eventdescription')
            return response
        else:
            return render_to_response("post_event.html", {'message': 'Insufficient data'}, context_instance=RequestContext(request))
    except:
        response = render_to_response("post_event.html", {
                                      'message': 'Something went to wrong'}, context_instance=RequestContext(request))
        response.delete_cookie('eventtitle')
        response.delete_cookie('startdate')
        response.delete_cookie('enddate')
        response.delete_cookie('plan')
        response.delete_cookie('category_name')
        response.delete_cookie('eventtype_name')
        response.delete_cookie('eventtype')
        response.delete_cookie('category')
        response.delete_cookie('eventdescription')
        return response


@csrf_exempt
def success(request):
    # field9 is payu success variable
    # this if condition for after success of payment
    if 'field9' in request.POST:
        response = render_to_response(
            "success.html", context_instance=RequestContext(request))
        response.delete_cookie('eventtitle')
        response.delete_cookie('startdate')
        response.delete_cookie('enddate')
        response.delete_cookie('plan')
        response.delete_cookie('category_name')
        response.delete_cookie('eventtype_name')
        response.delete_cookie('eventtype')
        response.delete_cookie('category')
        response.delete_cookie('eventdescription')
    else:
        response = HttpResponseRedirect('/')
    return response


def importcollegedata(request):

    saved = False
    saved_leads = []

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():

            f = request.FILES['file']
            path = os.path.join(settings.MEDIA_ROOT, 'imports')

            if not os.path.isdir(path):
                os.mkdir(path)

            path = os.path.join(path, "%s_%s" % (request.user.pk, time.time()))

            destination = open(path, 'wb+')
            for chunk in f.chunks():
                destination.write(chunk)
            destination.close()

            workbook = openpyxl.load_workbook(filename=path)
            sheet_names = workbook.get_sheet_names()

            for sheet_name in sheet_names:
                sheet = workbook.get_sheet_by_name(sheet_name)
                if sheet_name == 'Sheet1':
                    rows = sheet.get_highest_row()
                    cols = sheet.get_highest_column()

                    city_mapping = {}
                    collegeaddress_mapping = {}
                    collegecategory_mapping = {}

                    city_fields = [f.attname for f in City._meta.local_fields]
                    collegename_fields = [
                        f.attname for f in CampusCollege._meta.local_fields]
                    collegecategory_fields = [
                        f.attname for f in EventsCategory._meta.local_fields]

                    if rows > 1 and cols > 0 and sheet.cell(row=0, column=0) is not None:
                        for i in range(cols):
                            v = sheet.cell(row=0, column=i).value
                            v = v.replace(' ', '_').lower()

                            if v.startswith('city'):
                                v = v.replace('city_', '')
                                if v in city_fields:
                                    city_mapping[v] = i

                            if v.startswith('state'):
                                v = v.replace('state_', '')
                                if v in city_fields:
                                    city_mapping[v] = i

                            if v.startswith('country_code'):
                                v = v.replace('country_code_', '')
                                if v in city_fields:
                                    city_mapping[v] = i

                            if v.startswith('country_name'):
                                v = v.replace('country_name_', '')
                                if v in city_fields:
                                    city_mapping[v] = i

                            if v.startswith('college_name'):
                                v = v.replace('college_name_', '')
                                if v in collegename_fields:
                                    collegeaddress_mapping[v] = i

                            if v.startswith('name'):
                                v = v.replace('name_', '')
                                if v in collegecategory_fields:
                                    collegecategory_mapping[v] = i

                            else:
                                pass
                          # if v in lead_fields:
                          #   lead_mapping[v] = i

                        for i in range(1, rows):
                            citylist = City() if len(city_mapping.keys()) > 0 else None
                            collegelist = CampusCollege() if len(collegeaddress_mapping.keys()) > 0 else None
                            collegecategorylist = EventsCategory() if len(
                                collegecategory_mapping.keys()) > 0 else None

                            for field in city_mapping:
                                col = city_mapping[field]
                                v = sheet.cell(row=i, column=col)
                                if v:
                                    setattr(citylist, field, v.value)

                            for field in collegeaddress_mapping:
                                col = collegeaddress_mapping[field]
                                v = sheet.cell(row=i, column=col)
                                if v:
                                    setattr(collegelist, field, v.value)

                            for field in collegecategory_mapping:
                                col = collegecategory_mapping[field]
                                v = sheet.cell(row=i, column=col)
                                if v:
                                    setattr(collegecategorylist,
                                            field, v.value)

                            if citylist:
                                if not City.objects.filter(city=citylist.city).exists():
                                    citylist.save()

                            if collegecategorylist:
                                if not EventsCategory.objects.filter(category_name=collegecategorylist.name).exists():
                                    collegecategorylist.save()

                            if collegelist:
                                if City.objects.filter(city=citylist.city).exists():
                                    cityvalue = City.objects.filter(
                                        city=citylist.city)
                                    for c in cityvalue:
                                        collegelist.city_id = c.id
                                    collegelist.save()

                                if EventsCategory.objects.filter(category_name=collegecategorylist.name).exists():
                                    categoryvalue = EventsCategory.objects.filter(
                                        category_name=collegecategorylist.name)
                                    for c in categoryvalue:
                                        collegelist.collegetype_id = c.id
                                    collegelist.save()

            saved = True
            success_import = 'Successfully imported'
            return render_to_response('import.html', {'success_import': success_import, 'form': form, 'saved': saved, 'saved_leads': saved_leads}, context_instance=RequestContext(request))
    else:
        form = UploadFileForm()

    return render_to_response('import.html', {'form': form, 'saved': saved, 'saved_leads': saved_leads},
                              context_instance=RequestContext(request))


@csrf_exempt
def feedback(request):
    if request.is_ajax():
        if request.method == 'POST':
            print "enter"
            form = WebsiteFeedbackForm(request.POST)
            if form.is_valid():
                print "form is valid"
                # print "form",form
                form.save()
                print "save"
            else:
                form = WebsiteFeedbackForm()

            msg = "The operation has been received correctly."
    else:
        msg = "Fail"

    return HttpResponse(msg)


def getstate(request):
    from collections import OrderedDict
    results = []
    unsort_dict = {}
    key_loc = request.GET.get('term')
    state_lists = City.objects.filter(state__icontains=key_loc)

    for state_list in state_lists:
        statename = state_list.state.strip()
        stateid = state_list.id
        unsort_dict[statename] = {'stateid': stateid,
                                  'label': statename, 'value': statename}

    sorted_dic = OrderedDict(
        sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
    for k, v in sorted_dic.iteritems():
        results.append(v)

    return HttpResponse(simplejson.dumps(results), mimetype='application/json')

# admin side using by muthu
def getcollege(request):
    from collections import OrderedDict
    results = []
    unsort_dict = {}
    key_loc = request.GET.get('term')
    college_lists = CampusCollege.objects.filter(
        college_name__icontains=key_loc)
    for college_list in college_lists:
        collegename = college_list.college_name.strip()
        unsort_dict[collegename] = {'label': collegename, 'value': collegename}

    sorted_dic = OrderedDict(
        sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
    for k, v in sorted_dic.iteritems():
        results.append(v)
    print results
    return HttpResponse(simplejson.dumps(results), mimetype='application/json')

# admin side using by muthu
def getstate(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	state_lists = City.objects.filter(state__icontains=key_loc)

	for state_list in state_lists:
		statename = state_list.state.strip()
		stateid = state_list.id
		unsort_dict[statename] = {'stateid':stateid, 'label':statename, 'value':statename}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():
		results.append(v)



def getdept(request):
    from collections import OrderedDict
    results = []
    unsort_dict = {}
    key_loc = request.GET.get('term')
    college = request.GET.get('college')
    filterargs = {'college__id': college}
    # print CollegeDepartment.objects.filter(college__id=college)
    department_lists = CampusDepartment.objects.filter(
        department__icontains=key_loc)
    for department_list in department_lists:
        departmentname = department_list.department.strip()
        departmentid = department_list.id
        unsort_dict[departmentname] = {
            'departmentid': departmentid, 'label': departmentname, 'value': departmentname}

    sorted_dic = OrderedDict(
        sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
    for k, v in sorted_dic.iteritems():
        results.append(v)

    return HttpResponse(simplejson.dumps(results), mimetype='application/json')


def getcity_base(request):
    from collections import OrderedDict
    results = []
    unsort_dict = {}
    key_loc = request.GET.get('term')
    city_lists = City.objects.filter(city__icontains=key_loc)

    for city_list in city_lists:
        cityname = city_list.city.strip()
        cityid = city_list.id
        unsort_dict[cityname] = {'cityid': cityid,
                                 'label': cityname, 'value': cityname}

    sorted_dic = OrderedDict(
        sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
    for k, v in sorted_dic.iteritems():
        results.append(v)

    return HttpResponse(simplejson.dumps(results), mimetype='application/json')


def event_for_subcategory(request):
    if request.is_ajax() and request.GET and 'sub_category_id' in request.GET:
        objs1 = Postevent.objects.filter(festtype_id=sub_category_id)
        for obj in objs1:
            print obj.brand_name
        return JSONResponse([{'id': o1.id, 'name': smart_unicode(o1.brand_name)}
                             for o1 in objs1])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})


# To Load Categories in Home Page as Left Side bar
def all_subcategory_for_category(request):
    if request.is_ajax():
        objs1 = EventsSubCategory.objects.all()
        return JSONResponse([{'name': o1.subcategory_name, 'id': o1.id, 'icon': smart_unicode(o1.subcategory_icon), 'hover_icon': smart_unicode(o1.subcategory_hover_icon)} for o1 in objs1])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})

# To Load SubCategories in Home Page as Left Side bar


def subcategory_for_category(request):
    if request.is_ajax() and request.GET and 'category_id' in request.GET:
        objs1 = EventsSubCategory.objects.filter(
            category__id=request.GET['category_id'])
        return JSONResponse([{'name': o1.subcategory_name, 'id': o1.id, 'icon': smart_unicode(o1.subcategory_icon), 'hover_icon': smart_unicode(o1.subcategory_hover_icon)} for o1 in objs1])
    else:
        return JSONResponse({'error': 'No Ajax or No Get Request'})


def find_colleges(request):
    from django.utils.encoding import smart_unicode, force_unicode
    if request.is_ajax() and request.GET and 'city_id' in request.GET:
        objs = CampusCollege.objects.filter(city=request.GET['city_id'])
        return JSONResponse([{'id': o.id, 'name': smart_unicode(o.college_name)}
                             for o in objs])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})


def find_department(request):
    from django.utils.encoding import smart_unicode, force_unicode
    if request.is_ajax() and request.GET and 'college_id' in request.GET:
        objs = CampusDepartment.objects.filter(
            college__id=request.GET['college_id'])
        return JSONResponse([{'id': o.department.id, 'name': smart_unicode(o.department)}
                             for o in objs])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})


def find_subcategory(request):
    from django.utils.encoding import smart_unicode, force_unicode
    if request.is_ajax() and request.GET and 'category_id' in request.GET:
        objs = EventsSubCategory.objects.filter(
            category__id=request.GET['category_id'])
        return JSONResponse([{'id': o.id, 'name': smart_unicode(o.subcategory_name)}
                             for o in objs])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})


def find_city(request):
    from django.utils.encoding import smart_unicode, force_unicode
    if request.is_ajax() and request.GET and 'state' in request.GET:
        objs = City.objects.filter(state=request.GET['state'])
        return JSONResponse([{'id': o.id, 'name': smart_unicode(o.city)}
                             for o in objs])
    else:
        return JSONResponse({'error': 'Not Ajax or no GET'})


def getcity(request):
    from collections import OrderedDict
    results = []
    unsort_dict = {}
    key_loc = request.GET.get('term')
    state = request.GET.get('state')
    filterargs = {'state': state, 'city__icontains': key_loc}
    city_lists = City.objects.filter(**filterargs)
    for city_list in city_lists:
        cityname = city_list.city.strip()
        cityid = city_list.id
        unsort_dict[cityname] = {'cityid': cityid,
                                 'label': cityname, 'value': cityname}

    sorted_dic = OrderedDict(
        sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
    for k, v in sorted_dic.iteritems():
        results.append(v)

    return HttpResponse(simplejson.dumps(results), mimetype='application/json')


def home_v2(request):
    context = RequestContext(request,
                             {'request': request,
                              'user': request.user})
    return render_to_response('home_v2.html',
                              context_instance=context)


def get_events_for_calendar(request):
    import datetime
    events = Postevent.objects.all()
    time = datetime.time(10, 25)
    events_list = []
    for event in events:
        event_data = {'id': str(event.id), 'title': event.event_title, 'start': smart_unicode(datetime.datetime.combine(
            event.event_startdate_time, time)), 'end': smart_unicode(datetime.datetime.combine(event.event_enddate_time, time))}
        events_list.append(event_data)
    # print "event_list", events_list
    return HttpResponse(simplejson.dumps(events_list), mimetype='application/json')


def add_google_calendar(request, id=None):
    print "add_google_calendar"
    try:
        import argparse
        flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
    except ImportError:
        flags = None
    SCOPES = 'https://www.googleapis.com/auth/calendar.readonly'
    print "SCOPES", SCOPES
    CLIENT_SECRET_FILE = 'client_secret_google_calendar.json'
    print "CLIENT_SECRET_FILE", CLIENT_SECRET_FILE
    APPLICATION_NAME = 'Google Calendar API Python Quickstart'
    print "APPLICATION_NAME", APPLICATION_NAME

    def get_credentials():
        """Gets valid user credentials from storage.

        If nothing has been stored, or if the stored credentials are invalid,
        the OAuth2 flow is completed to obtain the new credentials.

        Returns:
                Credentials, the obtained credential.
        """
        home_dir = os.path.expanduser('~')
        credential_dir = os.path.join(home_dir, '.credentials')
        if not os.path.exists(credential_dir):
            os.makedirs(credential_dir)
        credential_path = os.path.join(credential_dir,
                                       'calendar-python-quickstart.json')

        store = oauth2client.file.Storage(credential_path)
        credentials = store.get()
        if not credentials or credentials.invalid:
            flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
            flow.user_agent = APPLICATION_NAME
            if flags:
                credentials = tools.run_flow(flow, store, flags)
            else:  # Needed only for compatibility with Python 2.6
                credentials = tools.run(flow, store)
            print('Storing credentials to ' + credential_path)
        return credentials

    def main():
        """Shows basic usage of the Google Calendar API.

        Creates a Google Calendar API service object and outputs a list of the next
        10 events on the user's calendar.
        """
        credentials = get_credentials()
        http = credentials.authorize(httplib2.Http())
        service = discovery.build('calendar', 'v3', http=http)

        now = datetime.datetime.utcnow().isoformat() + 'Z'  # 'Z' indicates UTC time
        print('Getting the upcoming 10 events')
        eventsResult = service.events().list(
            calendarId='primary', timeMin=now, maxResults=10, singleEvents=True,
            orderBy='startTime').execute()
        events = eventsResult.get('items', [])

        if not events:
            print('No upcoming events found.')
        for event in events:
            start = event['start'].get('dateTime', event['start'].get('date'))
            print(start, event['summary'])

    if __name__ == '__main__':
        main()

# gmail login redirect view code by priya

from django.core.urlresolvers import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.base import View
from social_auth.views import complete


class AuthComplete(View):

    def get(self, request, *args, **kwargs):
        backend = kwargs.pop('backend')
        try:
            return complete(request, backend, *args, **kwargs)
        except:
            messages.error(
                request, "Your Google Apps domain isn't authorized for this app")
            return HttpResponseRedirect(reverse('login'))


class LoginError(View):

    def get(self, request, *args, **kwargs):
        return HttpResponse(status=401)
